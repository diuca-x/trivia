from flask import Blueprint, jsonify

from flask_restful import Api

from api.resources.routes import SingleQuestion, specificQuestion, addOptions, handleOption, startGame,asd

#--- for the excel
import openpyxl
from flask import render_template, request
from fileinput import filename
import requests
import os

blueprint = Blueprint("api", __name__, url_prefix="/api")
api = Api(blueprint, errors=blueprint.errorhandler)

api.add_resource(SingleQuestion, "/question")
api.add_resource(specificQuestion, "/question/<int:question_id>")
api.add_resource(addOptions, "/addOptions/<int:question_id>")
api.add_resource(handleOption, "/handleOption/<int:option_id>")
api.add_resource(startGame, "/startGame/<int:ammount>" )
api.add_resource(asd, "/asd" )

@blueprint.route("/excel")
def excel_loader(): 
    return render_template("file_loader.html")

@blueprint.route("/load_file", methods = ["POST"])
def file_loader():
    file = request.files['file']
    file.save(file.filename)
    wb = openpyxl.load_workbook(file.filename)
    ws = wb.active
    
    has_text = "a"
    row_ammount = 0
    c = 2


    while (has_text):
        has_text = ws.cell(row=c, column=1).value
        
        if(has_text):
            row_ammount += 1
        c += 1

    
    for i in range(2,row_ammount + 2):
        
        question_to_add = {
                "question": "",
                "answer": "",
                "options" : []
            }
   
        question_to_add["question"] = ws.cell(row=i,column=1).value  
            
        question_to_add["answer"] = ws.cell(row=i,column=2).value
                
            
        options = ws.cell(row=i,column=3).value.split(",")

        question_to_add["options"] = options
    
        response = requests.post(os.environ.get("VITE_BACKEND_URL", False) + "api/question", json=question_to_add)

        print(response.json())
    


    return jsonify({"msg": "asd"})