from flask import Blueprint, jsonify, make_response, redirect, url_for


from flask_restful import Api

from api.resources.routes import SingleQuestion, specificQuestion, addOptions, handleOption, startGame,asd
from api.admin_routes import Signupator, Loginator

#--- for the excel
import openpyxl
from flask import render_template, request
from fileinput import filename
import requests
import os

from flask_jwt_extended import jwt_required




blueprint = Blueprint("api", __name__, url_prefix="/api")
auth_blueprint = Blueprint("auth", __name__, url_prefix="/auth")

api = Api(blueprint, errors=blueprint.errorhandler)
auth =  Api(auth_blueprint, errors=auth_blueprint.errorhandler)

api.add_resource(SingleQuestion, "/question")
api.add_resource(specificQuestion, "/question/<int:question_id>")
api.add_resource(addOptions, "/addOptions/<int:question_id>")
api.add_resource(handleOption, "/handleOption/<int:option_id>")
api.add_resource(startGame, "/startGame/<int:ammount>" )
api.add_resource(asd, "/asd" )

auth.add_resource(Signupator, "/signupator")
auth.add_resource(Loginator, "/loginator")

@auth_blueprint.route("/login")
def login(): 
   return render_template("login.html")


@auth_blueprint.route("/loged")
def loged(): 
   return render_template("loged.html")

@auth_blueprint.route("/signup")
def signup(): 
     return render_template("signup.html")





@blueprint.route("/excel")
def excel_loader(): 
    return render_template("file_loader.html")

@blueprint.route("/load_file", methods = ["POST"])
@jwt_required()
def file_loader():
    print(request.files)
    file = request.files['file']
    file.save(file.filename)
    wb = openpyxl.load_workbook(file.filename)
    ws = wb.active
    
    has_text = "a"
    row_ammount = 0
    c = 2


    while (has_text):
        has_text = ws.cell(row=c, column=1).value
        
        if(has_text):
            row_ammount += 1
        c += 1

    if(row_ammount) < 10:
        return make_response(jsonify({"msg" : "Error, less than 10 questions"}),400)
    
    for i in range(2,row_ammount + 2):
        
        question_to_add = {
                "question": "",
                "answer": "",
                "options" : []
            }
   
        question_to_add["question"] = ws.cell(row=i,column=1).value  
            
        question_to_add["answer"] = ws.cell(row=i,column=2).value
                
            
        options = ws.cell(row=i,column=3).value.split(",")

        question_to_add["options"] = options
        question_to_add["date"] = ws.cell(row=i,column=4).value.strftime('%d/%m/%Y')
        
        response = requests.post(os.environ.get("VITE_BACKEND_URL", False) + "api/question", json=question_to_add)
        
        print(response.json())
    


    return jsonify({"msg": "added"})